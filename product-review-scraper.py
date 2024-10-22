import os
import time
from typing import List, Optional
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, HttpUrl
from curl_cffi import requests as cureq
from curl_cffi.requests.errors import RequestsError

# Pydantic models to define structure for JSON parsing
# check the json structure in response via network lab
class Review(BaseModel):
    id: int
    childId: int
    title: str
    description: str
    name: str
    createdOn: str
    reviewCreationText: str
    likeCount: int
    rating: int
    isLikedByUser: bool
    isBuyer: bool
    images: Optional[List[HttpUrl]]

class ResponseData(BaseModel):
    page: int
    reviewData: List[Review]

class ApiResponse(BaseModel):
    response: ResponseData

# Function to remove duplicate reviews based on description and rating
def remove_duplicate_reviews(reviews):
    seen = set()
    unique_reviews = []
    for review in reviews:
        review_key = (review[1], review[2])  # (description, rating) as the unique key
        if review_key not in seen:
            seen.add(review_key)
            unique_reviews.append(review)
    return unique_reviews

# Main function to scrape product reviews and write them to an Excel sheet

def scrape_product_reviews(product_id, product_name, product_price, category, workbook, max_retries=3, timeout=60):
    reviews_data = []
    sort_categories = ['MOST_RECENT', 'MOST_USEFUL', 'MOST_HELPFUL', 'POSITIVE_FIRST', 'NEGATIVE_FIRST']

    # Iterate through sorting types and pages
    for sort_type in sort_categories:
        for page_no in range(1, 41):
            #upload the url of the api here
            url = f"https://____________/gateway-api/products/{product_id}/reviews?pageNo={page_no}&sort={sort_type}&filters=DEFAULT&domain=nykaa"

            # Retry logic for network requests
            for attempt in range(max_retries):
                try:
                    # Make API request
                    resp = cureq.get(url, impersonate="chrome", timeout=timeout)
                    print(f"Sort: {sort_type}, Page {page_no} - Status Code: {resp.status_code}")

                    if resp.status_code != 200:
                        print(f"Error: Received status code {resp.status_code} for page {page_no}")
                        continue

                    # Parse the API response using Pydantic
                    data = resp.json()
                    parsed_data = ApiResponse(**data)

                    # Collect reviews (without the reviewer name)
                    for review in parsed_data.response.reviewData:
                        reviews_data.append([sort_type, review.description, review.rating])

                    # Exit retry loop if request succeeds
                    break

                except RequestsError as e:
                    print(f"Attempt {attempt + 1}/{max_retries} failed: {e}")
                    if attempt < max_retries - 1:
                        print("Retrying in 5 seconds...")
                        time.sleep(5)
                    else:
                        print(f"Max retries reached for sort {sort_type} on page {page_no}")
                        break

    # Remove duplicate reviews
    reviews_data = remove_duplicate_reviews(reviews_data)

    # Write data to Excel sheet
    ws = workbook.active
    for row in reviews_data:
        ws.append([category, product_name, product_price] + row)

    print(f"Data for product with ID {product_id} added under category '{category}'.")

# Load or create Excel file with product data
def load_input_file(input_file_name):
    if not os.path.exists(input_file_name):
        raise FileNotFoundError(f"Input file '{input_file_name}' does not exist.")
    return load_workbook(input_file_name)

# Read products from the input sheet
def read_products_from_sheet(sheet):
    products = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        url, product_id, product_name, product_price, category = row
        products.append({"url": url, "id": product_id, "product_name": product_name, "product_price": product_price, "category": category})
    return products

# Main workflow function to process all categories and save output
def process_all_categories(input_file_name, output_file_name):
    input_wb = load_input_file(input_file_name)

    # Create a new Excel workbook for storing scraped reviews
    output_wb = Workbook()
    ws = output_wb.active
    ws.title = "All Reviews"
    headers = ["Category", "Product Name", "Product Price", "Sort Type", "Reviews", "Rating"]
    ws.append(headers)

    # Iterate over each sheet in the input workbook
    for sheet_name in input_wb.sheetnames:
        sheet = input_wb[sheet_name]
        products_to_scrape = read_products_from_sheet(sheet)

        # Scrape reviews for each product
        for product in products_to_scrape:
            scrape_product_reviews(product["id"], product["product_name"], product["product_price"], product["category"], output_wb)

    # Save output workbook
    output_wb.save(output_file_name)
    print(f"All data has been written to '{output_file_name}'")

if __name__ == "__main__":
    # File paths for input and output Excel files
    input_file_name = "./url of product/product_categoriesinsheet.xlsx"
    output_file_name = r"C:\Users\niraj\Downloads\scrapingusingapi\.venv\Lib\data of product\AllReviews.xlsx"

    try:
        process_all_categories(input_file_name, output_file_name)
        print("All categories have been processed successfully.")
    except FileNotFoundError as e:
        print(e)
    except Exception as e:
        print(f"An error occurred: {e}")
