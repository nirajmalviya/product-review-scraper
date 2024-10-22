import time
import openpyxl
import random
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from urllib.parse import urlparse, parse_qs

# Function to get a random user-agent to mimic different browser requests and avoid detection
def get_random_user_agent():
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.63 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.2 Safari/605.1.15",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36",
    ]
    return random.choice(user_agents)

# Initialize the Excel workbook to store the scraped data
workbook = openpyxl.Workbook()

# Dictionary mapping product categories to their URLs (replace these URLs with actual ones from the target website)
categories = {
    "Face Wash": "https://www.example.com/skin/cleansers/face-wash/c/8379?page_no={}&sort=popularity&eq=desktop",
    "Body Lotions": "https://www.example.com/bath-body/body-care/body-lotions-moisturizers/c/371?page_no={}&sort=popularity&eq=desktop",
    "Moisturizers": "https://www.example.com/skin/moisturizers/face-moisturizer-day-cream/c/8394?page_no={}&sort=popularity&eq=desktop"
}

# Loop through each category and its corresponding URL
for category_name, base_url in categories.items():
    # Create a new sheet in the Excel workbook for each category
    sheet = workbook.create_sheet(title=category_name)
    # Add headers for the sheet
    sheet.append(["url", "id", "product_name", "product_price", "category"])

    # Loop through pages (1 to 5) for each category to get multiple products
    for page_no in range(1, 6):
        # Set up Chrome options to add a random user-agent for each request
        chrome_options = Options()
        chrome_options.add_argument(f"user-agent={get_random_user_agent()}")

        # Initialize the Selenium WebDriver with the configured Chrome options
        driver = webdriver.Chrome(options=chrome_options)

        # Format the URL with the current page number
        url = base_url.format(page_no)
        driver.get(url)

        # Wait for the page to load with a random delay to avoid detection
        time.sleep(random.randint(5, 10))

        try:
            # Scrape all product URLs on the page
            product_urls = driver.find_elements(By.CLASS_NAME, "css-qlopj4")  # Update this class if necessary

            # Loop through each product URL and extract relevant information
            for product in product_urls:
                product_url = product.get_attribute('href')

                # Extract the product ID from the URL using query parameters or the URL path
                parsed_url = urlparse(product_url)
                product_id = None

                # Try to extract 'productId' from the query string, or fall back to the URL path
                if 'productId' in parse_qs(parsed_url.query):
                    product_id = parse_qs(parsed_url.query).get('productId', [''])[0]
                else:
                    # Extract product ID from the URL path if not present in the query string
                    path_parts = parsed_url.path.split('/')
                    for part in path_parts:
                        if part.isdigit():  # Assuming the product ID is numeric
                            product_id = part
                            break

                # Extract product name
                try:
                    product_name_element = product.find_element(By.CLASS_NAME, "css-xrzmfa")  # Update this class if necessary
                    product_name = product_name_element.text
                except:
                    product_name = "N/A"  # Handle missing product name

                # Extract product price
                try:
                    product_price_element = product.find_element(By.CLASS_NAME, "css-111z9ua")  # Update this class if necessary
                    product_price = product_price_element.text
                except:
                    product_price = "N/A"  # Handle missing product price

                # Append the product data to the Excel sheet
                sheet.append([product_url, product_id, product_name, product_price, category_name])

        except Exception as e:
            # Handle any errors during the scraping process
            print(f"Error on page {page_no} of {category_name}: {e}")

        # Close the browser after processing the current page
        driver.quit()

        # Wait for a random interval before moving to the next page to simulate human-like behavior
        time.sleep(random.randint(10, 20))

# Remove the default first sheet (if it's still present)
if "Sheet" in workbook.sheetnames:
    workbook.remove(workbook["Sheet"])

# Save the final Excel file to the specified directory
file_path = r"C:\Users\niraj\Downloads\scrapingusingapi\.venv\Lib\url of product\product_categoriesinsheet.xlsx"
workbook.save(file_path)

# Notify the user that the data has been saved successfully
print(f"Data saved successfully to {file_path}")
